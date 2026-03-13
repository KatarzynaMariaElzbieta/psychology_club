import csv
import os
from typing import Iterable, List

from email_validator import EmailNotValidError, validate_email
from openpyxl import load_workbook


EMAIL_HEADERS = {"email", "e-mail", "mail", "adres email", "adres e-mail"}


def _normalize_header(value: str) -> str:
    return (value or "").strip().lower()


def _extract_emails_from_rows(rows: Iterable[Iterable[str]], header: List[str] | None) -> List[str]:
    email_index = 0
    if header:
        normalized = [_normalize_header(col) for col in header]
        for idx, name in enumerate(normalized):
            if name in EMAIL_HEADERS or "email" in name:
                email_index = idx
                break

    seen = set()
    emails: List[str] = []
    invalid_count = 0
    invalid_samples: List[str] = []

    for row in rows:
        if not row:
            continue
        value = row[email_index] if email_index < len(row) else ""
        candidate = str(value or "").strip()
        if not candidate:
            continue
        try:
            normalized = validate_email(candidate, check_deliverability=False).email
        except EmailNotValidError:
            invalid_count += 1
            if len(invalid_samples) < 5:
                invalid_samples.append(candidate)
            continue
        if normalized not in seen:
            seen.add(normalized)
            emails.append(normalized)

    if invalid_count:
        samples = ", ".join(invalid_samples)
        suffix = f" (np. {samples})" if samples else ""
        raise ValueError(f"Znaleziono {invalid_count} nieprawidłowych adresów e-mail{suffix}")
    return emails


def parse_recipient_file(file_path: str) -> List[str]:
    _, ext = os.path.splitext(file_path)
    ext = ext.lower().strip()

    if ext == ".csv":
        return _parse_csv(file_path)
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _parse_xlsx(file_path)

    raise ValueError("Obsługiwane są tylko pliki CSV lub XLSX")


def _parse_csv(file_path: str) -> List[str]:
    encodings = ["utf-8-sig", "utf-8", "cp1250", "latin-1"]
    last_error: Exception | None = None

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                rows = list(reader)
            header = rows[0] if rows else []
            body = rows[1:] if rows else []
            return _extract_emails_from_rows(body, header)
        except Exception as exc:
            last_error = exc
            continue

    raise ValueError(f"Nie udało się odczytać pliku CSV: {last_error}")


def _parse_xlsx(file_path: str) -> List[str]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(value or "") for value in rows[0]]
    body = [list(row) for row in rows[1:]]
    return _extract_emails_from_rows(body, header)
